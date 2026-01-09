import openpyxl
import pandas as pd
from .util import (
    log,
    match,
    _arg_to_list,
    )


def hide(
        filename,
        axis='col',  #'row', 'col', 'sheet'
        patterns=None,
        regex=True,
        hide=True,  #True=hide, False=unhide
        sheet=None,
        verbosity=3,
        ):  #pragma: no cover (does not affect reading of xlsx files)
    """
    Hide or unhide columns, rows, or sheets in an Excel file.
    """

    if hide is True:
        mode = 'hidden'
    else:
        mode = 'visible'

    wb = openpyxl.load_workbook(filename)

    if axis in ['col', 'column', 'cols', 'columns', 1]:
        for ws in wb.worksheets:
            if sheet is None or ws.title == sheet:
                hidden = []
                for col in ws.columns:
                    if match(patterns, col[0].value, regex):
                        ws.column_dimensions[col[0].column_letter].hidden = hide
                        hidden.append(col[0].value)
                if hidden:
                    hidden = '\n'.join(hidden)
                    text = (
                        f'debug: columns made {mode} in'
                        f' "{filename}" sheet "{ws.title}":\n{hidden}'
                        )
                    log(text, 'qp.hide()', verbosity)


    elif axis in ['row', 'rows', 'index', 0]:
        for ws in wb.worksheets:
            if sheet is None or ws.title == sheet:
                hidden = []
                for row in ws.rows:
                    if match(patterns, row[0].value, regex):
                        ws.row_dimensions[row[0].row].hidden = hide
                        hidden.append(row[0].value)
                if hidden:
                    hidden = '\n'.join(hidden)
                    text = (
                        f'debug: rows made {mode} in'
                        f' "{filename}" sheet "{ws.title}":\n{hidden}'
                        )
                    log(text, 'qp.hide()', verbosity)


    elif axis in ['sheet', 'worksheet', 'tab', 2]:
        hidden = []
        for ws in wb.worksheets:
            if match(patterns, ws.title, regex):
                if hide:
                    ws.sheet_state = mode
                else:
                    ws.sheet_state = mode
                hidden.append(ws.title)
        if hidden:
            hidden = '\n'.join(hidden)
            text = f'debug: sheets made {mode} in "{filename}":\n{hidden}'
            log(text, 'qp.hide()', verbosity)

    else:
        log(f'error: unknown axis "{axis}"', 'qp.hide', verbosity)

    wb.save(filename)
    wb.close()


def format_excel(
        filename,
        sheet=None,
        freeze_panes='B2',
        hide_sheets=None,
        hide_cols=None,
        col_width_max=70,
        col_width_padding=2,
        align_vertical='top',
        align_horizontal='left',
        wrap_text=True,
        verbosity=3,
        ):  #pragma: no cover (does not affect reading of xlsx files)
    """
    applies formatting to an Excel file:
    - adjust col width to max length of cell content (accounts for linebreaks)
    - set cell alignment to top-left and wrap text
    - hide specified columns
    """

    wb = openpyxl.load_workbook(filename)
    if sheet:
        sheetnames = _arg_to_list(sheet)
    else:
        sheetnames = [sheetname for sheetname in wb.sheetnames]
    sheets_hide = _arg_to_list(hide_sheets)
    cols_hide = _arg_to_list(hide_cols)

    for sheetname in sheetnames:
        if sheetname in sheets_hide:
            continue

        data = pd.read_excel(filename, sheet_name=sheetname)
        sheet = wb[sheetname]

        if freeze_panes:
            sheet.freeze_panes = freeze_panes

        #adjust column widths, cell alignment and text wrapping
        for col in sheet.columns:
            if col[0].value is None:
                text = (
                    'warning: skipping column with'
                    f' no header in sheet "{sheet.title}"'
                    )
                log(text, 'qp.format_excel()', verbosity)
                continue
            if col[0].value in cols_hide:
                sheet.column_dimensions[col[0].column_letter].hidden = True
                continue

            colname = col[0].value
            col_letter = col[0].column_letter

            #multiline cells are split by newline,
            #expanded into new rows,
            #and the maximum length is calculated
            #these changes are not applied to the actual data,
            #only the column width is adjusted
            max_length = (
                data[colname]
                .astype(str)
                .apply(lambda x: x.split('\n'))
                .explode()
                .str
                .len()
                .max()
                )
            max_length = max(max_length, len(colname))
            max_length = min(max_length, col_width_max) + col_width_padding

            sheet.column_dimensions[col_letter].width = max_length

            for cell in col:
                cell.alignment = openpyxl.styles.Alignment(
                    vertical=align_vertical,
                    horizontal=align_horizontal,
                    wrap_text=wrap_text
                    )

    wb.save(filename)
    wb.close()
